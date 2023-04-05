#Importing the Libraries
import findspark
findspark.init()
import pyspark
from pyspark.sql import SparkSession 
from pyspark.sql.functions import when, col, concat, lit, sum, lpad, length
from pyspark.sql.window import Window
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

# Creating Spark Session
spark = SparkSession.builder.appName('DE I Take Home Assignment').getOrCreate()

# Reading /loading the Dataset from CSV file
enrollment_df = spark.read.load("C:/Users/Learner_XZHCG216/Desktop/DE I Take Home Assignment/postsecondary_enrollment_current_2020-21.csv", format="csv", header=True, inferSchema=True)

# Rename columns in enrollment
enrollment_df = enrollment_df.withColumnRenamed('DISTRICT_NAME', 'district_name').withColumnRenamed('SCHOOL_NAME', 'school_name')

# Replace null values with default value 9999 in DISTRICT_CODE and SCHOOL_CODE columns for completion_df
enrollment_df = enrollment_df.fillna({'DISTRICT_CODE': '9999', 'SCHOOL_CODE': '9999'})
    
# Standardize the DISTRICT_CODE and SCHOOL_CODE for completion_df
enrollment_df = enrollment_df.withColumn(
    "DISTRICT_CODE", lpad(col("DISTRICT_CODE"), 4, '0')
).withColumn(
    "SCHOOL_CODE", lpad(col("SCHOOL_CODE"), 4, '0')
)

# Add entity_type column to enrollment_df based on SCHOOL_NAME
enrollment_df = enrollment_df.withColumn(
    "entity_type",
    when(col("SCHOOL_NAME") == "[Statewide]", "state")
    .when(col("SCHOOL_NAME") == "[Districtwide]", "district")
    .otherwise("school")
)

# Add state_id column to enrollment_df based on SCHOOL_NAME
enrollment_df = enrollment_df.withColumn(
    "state_id",
    when(col("SCHOOL_NAME") == "[Statewide]", "state")
    .when(col("SCHOOL_NAME") == "[Districtwide]", col("DISTRICT_CODE"))
    .otherwise(concat(col("DISTRICT_CODE"), col("SCHOOL_CODE")))
)

# Filter enrollment_df by INITIAL_ENROLLMENT column
enrollment_df = enrollment_df.filter(
    (col("INITIAL_ENROLLMENT") == "First Fall") &
    (~col("INITIAL_ENROLLMENT").isin(["Later Enrollment", "Second Fall", "*"]))
)

# Filter GROUP_BY_VALUE column
group_values = ["All Students", "Amer Indian", "Asian", "Black", "Econ Disadv", "ELL/LEP", "Eng Prof (Non-LEP)", "Female", "Hispanic", "Male", "Not Econ Disadv", "Not Migrant", "Pacific Isle", "SwD", "SwoD (students without disabilities)", "Two or More (two or more races)", "White", "Migrant"]
enrollment_df = enrollment_df.filter(col("GROUP_BY_VALUE").isin(group_values) & ~(col("GROUP_BY_VALUE").isin(["[Data Suppressed]", "Unknown"])))

# Reading /loading the Dataset from CSV file
completion_df = spark.read.load("C:/Users/Learner_XZHCG216/Desktop/DE I Take Home Assignment/hs_completion_certified_2020-21.csv", format="csv", header=True, inferSchema=True)

# Remove whitespace from "DISTRICT_CODE " column name
completion_df = completion_df.withColumnRenamed('DISTRICT_CODE ', 'DISTRICT_CODE')

# Replace null values with default value 9999 in DISTRICT_CODE and SCHOOL_CODE columns for completion_df
completion_df = completion_df.fillna({'DISTRICT_CODE': '9999', 'SCHOOL_CODE': '9999'})

# Standardize the DISTRICT_CODE and SCHOOL_CODE for completion_df
completion_df = completion_df.withColumn(
    "DISTRICT_CODE", lpad(col("DISTRICT_CODE"), 4, '0')
).withColumn(
    "SCHOOL_CODE", lpad(col("SCHOOL_CODE"), 4, '0')
)

# Filter columns TIMEFRAME, COMPLETION_STATUS, COHORT
completion_df = completion_df.filter((col("TIMEFRAME") == "4-Year rate") &
                                     ~(col("TIMEFRAME").isin(["5-Year rate", "6-Year rate", "7-Year rate"])) &
                                     col("COMPLETION_STATUS").isin(["Completed - Regular High School Diploma", 
                                    "Completed - High School Equivalency Diploma",
                                    "Completed - Other High School Completion Credential"]) &
                                     ~(col("COMPLETION_STATUS").isin(["Not Completed - Continuing Toward Completion", 
                                    "Not Completed - Not Continuing",
                                    "Not Completed - Not Known to be Continuing Toward Completion",
                                    "Not Completed - Reached Maximum Age",
                                    "*"])) &
                                     (col("COHORT") == "2021") &
                                     ~(col("COHORT").isin(["2018", "2019", "2020"]))
) 

# Joining the two dataframes and ordering by GROUP_BY_VALUE
joined_df = enrollment_df.join(completion_df, ['SCHOOL_YEAR', 'DISTRICT_CODE', 'SCHOOL_CODE', 'GROUP_BY_VALUE'], 'inner') \
                         .orderBy("GROUP_BY_VALUE")

# Define window specification
window_spec = Window.partitionBy("SCHOOL_NAME", "GROUP_BY_VALUE")

# Add value_numerator column to enrollment_df
enrollment_df = enrollment_df.withColumn("value_numerator", sum("STUDENT_COUNT").over(window_spec).cast("int"))
# Add value_denominator column to completion_df
completion_df = completion_df.withColumn("value_denominator", sum("STUDENT_COUNT").over(window_spec).cast("int"))

# Joining the two dataframes
joined_df = enrollment_df.join(completion_df, ['SCHOOL_YEAR', 'DISTRICT_CODE', 'SCHOOL_CODE', 'GROUP_BY_VALUE'], 'inner') 

# Add value column to joined_df
joined_df = joined_df.withColumn("value", (col("value_numerator") / col("value_denominator")) * 100)

# Filter rows where value column contains "*"
joined_df = joined_df.filter(col("value").rlike("^[^*]+$"))

# Filter rows where value_denominator is 0
joined_df = joined_df.filter(col("value_denominator") != 0)

# Ordering by GROUP_BY_VALUE
joined_df = joined_df.orderBy("GROUP_BY_VALUE")

# Convert joined_df to a pandas dataframe
df = joined_df.toPandas()

# Add new columns with the same data
df['year'] = '2021'
df['state'] = 'WI'
df['data_type'] = 'Percent Enrolled in College Immediately Following High School'
df['file_name'] = 'postsecondary_enrollment_current_2020-21.csv'

# Group by GROUP_BY_VALUE and select the first row for each group
df = df.groupby(['GROUP_BY_VALUE', 'value_numerator', 'value_denominator', 'value']).first().reset_index()

# Rename columns
df = df.rename(columns={'GROUP_BY_VALUE': 'breakdown'})

# Reorder columns
df = df[['year', 'state', 'entity_type', 'state_id', 'district_name', 'school_name', 'data_type', 'breakdown', 'value_numerator', 'value_denominator', 'value', 'file_name']]

# Sort by entity_type in descending order first and then by matching school_name in ascending order
df = df.sort_values(by=['entity_type', 'district_name', 'school_name'], ascending=[False, True, True])

# Create a workbook object and get the active sheet
workbook = Workbook()
sheet = workbook.active

# Set up some style constants
HEADER_FILL = PatternFill(start_color='6A99D0', end_color='6A99D0', fill_type='solid')
ROW_FILL_1 = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
ROW_FILL_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Define font style for bold text
bold_font = Font(bold=True)

# Write the headers and apply the custom style and border
headers = df.columns.tolist()
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num, value=str(header))
    cell.fill = HEADER_FILL
    cell.font = bold_font

# Write the data and apply the custom style
for row_num, row_data in enumerate(df.values.tolist(), 2):
    row_style = ROW_FILL_1 if row_num % 2 == 0 else ROW_FILL_2
    for col_num, cell_value in enumerate(row_data, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=str(cell_value))
        cell.fill = row_style

# Save the workbook to disk
workbook.save('output.xlsx')
